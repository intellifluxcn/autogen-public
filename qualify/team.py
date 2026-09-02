"""Qualify stage team.

Runs Claude Code non-interactively (`claude -p`) against each downloaded dataset
unit and records a strict/loose/fail/error verdict in
``artifacts.qualification_status``. Work is strictly DB-driven (per-project
dataset artifacts), datasets are treated as untrusted (symlink/`..`/absolute
rejection + ext/size allowlist), and the subprocess is killed by process group
on timeout — mirroring ``download/local_cdp_runtime.py``.
"""

from __future__ import annotations

import asyncio
import contextlib
import gzip
import json
import logging
import zlib
import os
import re
import shutil
import signal
from collections import OrderedDict
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from orchestrator_ui import push_message, update_progress, update_stage
from utils.pipeline_log import pipeline_log

logger = logging.getLogger(__name__)

# Trust-boundary allowlist for files copied into the claude sandbox.
QUALIFY_ALLOWED_EXTENSIONS = frozenset(
    # .soft = GEO SOFT (GDS/GSE) — a text format carrying the expression data
    # table + metadata; common GEO download form that was previously rejected
    # (→ "no files materialized" error on .soft.gz units like GDS1761).
    {".csv", ".tsv", ".txt", ".json", ".gct", ".tab", ".md", ".soft"}
)
QUALIFY_MAX_FILE_BYTES = int(os.getenv("QUALIFY_MAX_FILE_BYTES", str(50 * 1024 * 1024)))
# Archive (.zip/.tar.gz/.tgz) extraction caps — datasets like GDSC/PharmacoGx
# ship their expression + drug tables inside archives. Members are filtered by
# the same allowlist; these bound the decompression-bomb / disk blast radius.
QUALIFY_ARCHIVE_MAX_TOTAL_BYTES = int(
    os.getenv("QUALIFY_ARCHIVE_MAX_TOTAL_BYTES", str(200 * 1024 * 1024))
)
QUALIFY_ARCHIVE_MAX_FILES = int(os.getenv("QUALIFY_ARCHIVE_MAX_FILES", "200"))

# claude -p runner.
QUALIFY_CLAUDE_BIN = os.getenv("QUALIFY_CLAUDE_BIN", "claude")
# Qualify judge model ladder: run the cheaper/faster PRIMARY model first, and
# auto-escalate to the stronger ESCALATION model only when the primary cannot
# produce a usable verdict (error / unparseable). Empty primary → claude CLI
# account default; empty escalation → no escalation. Accepts aliases or full IDs.
QUALIFY_CLAUDE_MODEL = os.getenv("QUALIFY_CLAUDE_MODEL", "sonnet").strip()
QUALIFY_CLAUDE_ESCALATION_MODEL = os.getenv("QUALIFY_CLAUDE_ESCALATION_MODEL", "opus").strip()


def _claude_argv(model: str, *extra: str) -> list:
    """Build the claude CLI argv, injecting --model when ``model`` is set."""
    argv = [QUALIFY_CLAUDE_BIN]
    if model:
        argv += ["--model", model]
    return argv + list(extra)


def _qualify_model_ladder() -> list:
    """Ordered, de-duplicated list of models to try (primary then escalation)."""
    ladder = [QUALIFY_CLAUDE_MODEL]
    if QUALIFY_CLAUDE_ESCALATION_MODEL and QUALIFY_CLAUDE_ESCALATION_MODEL != QUALIFY_CLAUDE_MODEL:
        ladder.append(QUALIFY_CLAUDE_ESCALATION_MODEL)
    return ladder
QUALIFY_CLAUDE_TIMEOUT_SECONDS = float(
    os.getenv("QUALIFY_CLAUDE_TIMEOUT_SECONDS", "600")
)
QUALIFY_HEALTHCHECK_TIMEOUT_SECONDS = float(
    os.getenv("QUALIFY_HEALTHCHECK_TIMEOUT_SECONDS", "30")
)
# Total claude attempts per dataset: a transient failure (spawn/communicate/
# non-zero exit — e.g. rate limit) is retried once. Timeouts and unparseable
# verdicts are NOT retried (expensive / deterministic).
QUALIFY_CLAUDE_MAX_ATTEMPTS = int(os.getenv("QUALIFY_CLAUDE_MAX_ATTEMPTS", "2"))
QUALIFY_GENE_OVERLAP_THRESHOLD = os.getenv("QUALIFY_GENE_OVERLAP_THRESHOLD", "0.70")

# 1.0 reference gene set (expression feature universe) for the strict gene-overlap
# test. Injected as a sandbox file (60k+ ENSG IDs — far too large to inline in the
# prompt), so claude can compute the overlap deterministically with shell tools.
# Missing/empty file → no injection → strict capped at loose (graceful fallback).
QUALIFY_REFERENCE_GENE_SET_PATH = os.getenv(
    "QUALIFY_REFERENCE_GENE_SET_PATH",
    str(Path(__file__).resolve().parent / "reference_genes.txt"),
)
_REFERENCE_GENES_SANDBOX_NAME = "__ai4med_reference_genes__.txt"

_VALID_VERDICTS = ("strict", "loose", "fail")
# `escalate` is a model self-assessment ("I can't confidently apply the
# contract") — the runner re-judges with the stronger model rather than
# recording it as a final verdict.
_COMMITTED_VERDICTS = ("strict", "loose", "fail")
# Anchored: the final non-empty line must be EXACTLY `VERDICT: <verdict>`.
_VERDICT_RE = re.compile(r"^VERDICT: (strict|loose|fail|escalate)$")
# Decisive reason, emitted on the line immediately before the VERDICT line.
_REASON_RE = re.compile(r"^REASON:\s*(.+)$")


def _empty_meta() -> Dict:
    return {
        "model": None,
        "input_tokens": 0,
        "output_tokens": 0,
        "cost_usd": 0.0,
        "duration_ms": 0,
    }


def _extract_meta(data: Dict, model_alias: str) -> Dict:
    """Pull model / token / cost / duration telemetry from the claude CLI
    ``--output-format json`` envelope. Input tokens sum fresh + cached so the
    figure reflects total tokens processed, not just the uncached delta."""
    usage = data.get("usage") or {}
    in_tok = (
        (usage.get("input_tokens") or 0)
        + (usage.get("cache_read_input_tokens") or 0)
        + (usage.get("cache_creation_input_tokens") or 0)
    )
    return {
        "model": model_alias or None,
        "input_tokens": in_tok,
        "output_tokens": usage.get("output_tokens") or 0,
        "cost_usd": data.get("total_cost_usd") or 0.0,
        "duration_ms": data.get("duration_ms") or 0,
    }


def _accumulate_meta(agg: Dict, meta: Dict) -> None:
    if not meta:
        return
    agg["input_tokens"] += meta.get("input_tokens") or 0
    agg["output_tokens"] += meta.get("output_tokens") or 0
    agg["cost_usd"] += meta.get("cost_usd") or 0.0
    agg["duration_ms"] += meta.get("duration_ms") or 0

_PROMPT_PATH = Path(__file__).resolve().parent / "validation_prompt.md"


class QualifyTeam:

    def __init__(
        self,
        team_name: str = "QualifyTeam",
        project_id: Optional[str] = None,
        *,
        storage_backend: Optional[str] = None,
    ):
        self.team_name = team_name
        self.project_id = project_id
        self.storage_backend = (
            storage_backend or os.getenv("STORAGE_BACKEND", "local")
        ).lower()
        self._dao = None

    # ----------------------------------------------------------------- helpers

    def _get_dao(self):
        if self._dao is None:
            from database.dao import ProjectDAO

            self._dao = ProjectDAO()
        return self._dao

    @staticmethod
    def parse_verdict(stdout: str) -> str:
        """Scan stdout from the END for the final ``^VERDICT: (strict|loose|fail)$``.

        The verdict line MUST be the final non-empty line (contract); anything
        after it (trailing chatter, a second line) means an untrusted/garbled
        run → ``error``. Missing/empty/non-matching → ``error``.
        """
        lines = stdout.splitlines()
        # Drop trailing whitespace-only lines (tolerated by the contract).
        i = len(lines) - 1
        while i >= 0 and lines[i].strip() == "":
            i -= 1
        if i < 0:
            return "error"
        match = _VERDICT_RE.match(lines[i])
        if match:
            return match.group(1)
        return "error"

    @staticmethod
    def parse_reason(text: str) -> str:
        """Extract the decisive one-line reason the model emits as ``REASON: ...``
        immediately before the VERDICT line. Scans from the end; returns the
        first match (truncated), or "" when absent."""
        for line in reversed(text.splitlines()):
            m = _REASON_RE.match(line.strip())
            if m:
                return m.group(1).strip()[:1000]
        return ""

    @staticmethod
    def group_dataset_units(rows: List[Dict]) -> "OrderedDict[str, List[Dict]]":
        """Group per-file dataset artifact rows into one unit per parent dir.

        A dataset artifact is a per-FILE row; a paper's dataset spans N rows that
        share a parent directory. Qualify each unit once; the verdict is written
        to every member row. Rows without a usable ``file_path`` are skipped.
        """
        groups: "OrderedDict[str, List[Dict]]" = OrderedDict()
        for row in rows:
            file_path = row.get("file_path")
            if not file_path:
                continue
            parent = os.path.dirname(file_path)
            groups.setdefault(parent, []).append(row)
        return groups

    # Injected for `_knowndb_*` units: canonical public pharmacogenomics DBs
    # (GDSC/CCLE/PRISM/CTRP/gCSI/GRAY/FIMM/UHNBreast/Tavor/…) fetched directly
    # from source are reference-grade. Their gene panel and drug library
    # legitimately differ from the ai4med 1.0 schema, so the 1.0-schema-matching
    # strict gates (gene-overlap, drug-subset) are waived — these ARE the
    # gold-standard linkable E+R. The `_knowndb_` prefix (not the DB name) is the
    # authoritative signal; the waiver text tells the judge not to re-gate on it.
    _KNOWN_DB_WAIVER = (
        "\nCANONICAL DB UNIT: The pipeline has ALREADY determined that THIS unit "
        "is a curated public pharmacogenomics database fetched directly from its "
        "source repository — treat that as GIVEN. Do NOT re-decide it from the "
        "database's name or from whether you recognize it: any unit carrying this "
        "notice qualifies (the set is open-ended — GDSC / CCLE / PRISM / CTRP / "
        "gCSI / GRAY / FIMM / UHNBreast / Tavor / … are examples, NOT an "
        "allowlist). Its expression and drug-response are reference-grade and "
        "linkable by stable cell-line/sample IDs (DepMap/ACH, COSMIC, or the "
        "library's own sample IDs). For THIS unit the two strict gates are WAIVED: "
        "assign \"strict\" if E + numeric R + L are all present, EVEN IF the "
        "gene-overlap is below {GENE_OVERLAP_THRESHOLD} or some/all drugs are "
        "outside the 1.0 compound list. (E+R+L still genuinely required; numeric-R "
        "and figure-extraction caps still apply.)"
        "\nExpression source: it is NORMAL and intended for a canonical unit to "
        "pair drug-response from one canonical library (e.g. GDSC/CTRP/gCSI/FIMM/"
        "UHNBreast) with expression reused from another canonical source (e.g. "
        "CCLE_expression) linked through the sample_info crosswalk — this is the "
        "designed pattern, NOT a disqualifying 'mixed-source' dataset. A "
        "self-contained unit that ships its OWN expression keyed by the same "
        "sample IDs as its R (no sample_info needed) is equally valid."
        "\nEstablishing L (linkage): "
        "(a) if E and R already share the same sample IDs, a non-empty direct "
        "intersection establishes L; "
        "(b) otherwise use the COMPLETE sample_info crosswalk that ships in the "
        "unit — L holds when R's cell-line identifiers (names or IDs), after "
        "normalization, resolve into sample_info, which maps them to the stable "
        "DepMap/ACH (and COSMIC) IDs that key the expression. "
        "The expression file may be a TRUNCATED HEAD SAMPLE of a large canonical "
        "matrix (filename prefixed \"__SAMPLE__\"): it still proves E is present "
        "and shows the axis labels. Whether an R-vs-E intersection is still valid "
        "depends on WHICH axis the sample IDs live on — check the header before "
        "deciding. (i) Samples are the ROWS (first column is the sample/cell-line "
        "ID, e.g. DepMap CCLE_expression): head truncation drops later sample "
        "rows, so an R-vs-(truncated E) row intersection is meaningless — use "
        "sample_info instead; a cell line absent from the truncated head but "
        "present in sample_info IS linkable. (ii) Samples are the COLUMNS (the "
        "HEADER row lists the sample IDs and the rows are genes, e.g. a "
        "genes×samples RNA-seq matrix): truncation drops only gene ROWS — EVERY "
        "sample column survives in the intact header line, so `head -1` yields the "
        "complete sample set and an R-vs-(E header columns) intersection IS a "
        "valid, complete linkage. For a self-contained unit with no sample_info, "
        "case (ii) is the intended way to establish L."
    )

    @classmethod
    def render_validation_prompt(
        cls, dataset_dir: str, *, reference_gene_count: int = 0, known_db: bool = False
    ) -> str:
        """Load the checked-in validation prompt template and inject the
        sandboxed dataset dir + spec defaults. When ``reference_gene_count`` is
        0 the reference-gene-set placeholder resolves to its empty fallback
        (strict unreachable, capped at loose); otherwise it points the model at
        the runner-provided reference file materialized in the sandbox."""
        text = _PROMPT_PATH.read_text(encoding="utf-8")
        template = cls._extract_prompt_template(text)
        return (
            template.replace("{DATASET_DIR}", dataset_dir)
            .replace("{EXTENDED_DRUG_ALLOWLIST}", "")
            .replace("{REFERENCE_GENE_SET}", cls._reference_gene_injection(reference_gene_count))
            .replace("{KNOWN_DB_CANONICAL}", cls._KNOWN_DB_WAIVER if known_db else "")
            .replace("{GENE_OVERLAP_THRESHOLD}", QUALIFY_GENE_OVERLAP_THRESHOLD)
        )

    @staticmethod
    def _reference_gene_injection(reference_gene_count: int) -> str:
        if reference_gene_count <= 0:
            return ""
        return (
            f"Provided as a newline-delimited file named "
            f"`{_REFERENCE_GENES_SANDBOX_NAME}` in the dataset directory "
            f"({reference_gene_count} Ensembl gene IDs). That file is "
            f"RUNNER-PROVIDED reference data (the 1.0 expression feature "
            f"universe), NOT part of the candidate dataset: exclude it from the "
            f"E/R/L detection and never treat it as the candidate's expression "
            f"or response matrix. Use it ONLY as reference_genes for the overlap "
            f"test — overlap_rate = (count of the candidate's gene IDs, mapped "
            f"to Ensembl where needed, that appear in this file) / "
            f"{reference_gene_count}. Use shell tools (e.g. grep -Fxf, comm, "
            f"wc -l) to count the intersection deterministically."
        )

    def _materialize_reference_genes(self, sandbox_dir: str) -> int:
        """Copy the 1.0 reference gene set into the sandbox under a reserved
        name so the strict gene-overlap test is computable. Returns the gene
        count, or 0 when no reference file is configured/available (strict then
        stays capped at loose)."""
        path = QUALIFY_REFERENCE_GENE_SET_PATH
        if not path or not os.path.isfile(path):
            return 0
        try:
            dest = os.path.join(sandbox_dir, _REFERENCE_GENES_SANDBOX_NAME)
            shutil.copyfile(path, dest)
            with open(dest, "r", encoding="utf-8") as fh:
                return sum(1 for line in fh if line.strip())
        except OSError as e:
            logger.warning("qualify: failed to materialize reference gene set: %s", e)
            return 0

    @staticmethod
    def _extract_prompt_template(markdown: str) -> str:
        """Pull the prompt body out of the first fenced ``` block; fall back to
        the whole document if no fence is present."""
        fence = "```"
        first = markdown.find(fence)
        if first == -1:
            return markdown
        start = markdown.find("\n", first)
        end = markdown.find(fence, start + 1)
        if start == -1 or end == -1:
            return markdown
        return markdown[start + 1 : end]

    @classmethod
    def _copy_into_sandbox(cls, src_dir: str, sandbox_dir: str) -> List[str]:
        """Copy allowlisted files from ``src_dir`` into ``sandbox_dir``,
        treating ``src_dir`` as hostile. Rejects symlinks, paths that escape
        ``src_dir`` (``..``), non-allowlisted extensions, and oversize files.
        Returns the list of sandbox file paths written."""
        src_root = Path(src_dir).resolve()
        sandbox = Path(sandbox_dir)
        sandbox.mkdir(parents=True, exist_ok=True)
        copied: List[str] = []
        if not src_root.is_dir():
            return copied
        for entry in sorted(src_root.iterdir()):
            name = entry.name
            # Reject symlinks outright (no following).
            if entry.is_symlink() or os.path.islink(str(entry)):
                continue
            if not entry.is_file():
                continue
            # Containment: the real resolved path must stay under src_root.
            try:
                resolved = entry.resolve()
                resolved.relative_to(src_root)
            except (ValueError, OSError):
                continue
            # Direct allowlisted file; a single-member .gz whose inner file is
            # allowlisted (GEO matrices ship as *.txt.gz); an .xlsx (converted);
            # or an archive (.zip/.tar.gz/.tgz/.tar) whose allowlisted members
            # are extracted. .gz is gunzipped with a decompressed-size cap.
            lname = name.lower()
            is_archive = (
                lname.endswith(".zip") or lname.endswith(".tar.gz")
                or lname.endswith(".tgz") or lname.endswith(".tar")
            )
            suffix = entry.suffix.lower()
            is_gz = suffix == ".gz" and not lname.endswith(".tar.gz")
            is_xlsx = suffix == ".xlsx"
            inner_name = name[:-3] if is_gz else name
            inner_suffix = Path(inner_name).suffix.lower()
            allowed_gz = is_gz and inner_suffix in QUALIFY_ALLOWED_EXTENSIONS
            if not (is_archive or suffix in QUALIFY_ALLOWED_EXTENSIONS or allowed_gz or is_xlsx):
                continue
            try:
                fsize = entry.stat().st_size
            except OSError:
                continue
            if is_archive:
                copied.extend(cls._extract_archive(str(entry), str(sandbox), entry.stem))
            elif is_xlsx:
                # Supplementary tables often ship as .xlsx — convert each sheet
                # to CSV in the sandbox (output is row/size-capped internally).
                copied.extend(cls._convert_xlsx(
                    str(entry), str(sandbox), entry.stem, QUALIFY_MAX_FILE_BYTES
                ))
            elif allowed_gz:
                # _safe_gunzip keeps a head sample when the decompressed stream
                # exceeds the cap (large GEO matrices) rather than dropping it.
                dest = sandbox / inner_name
                if cls._safe_gunzip(str(entry), str(dest), QUALIFY_MAX_FILE_BYTES):
                    copied.append(str(dest))
            elif fsize > QUALIFY_MAX_FILE_BYTES:
                # Oversize plain file (e.g. a multi-GB NCI-60/CCLE matrix): keep a
                # head sample so claude can still inspect axes / sample IDs.
                dest = sandbox / ("__SAMPLE__" + name)
                if cls._write_head_sample(str(entry), str(dest), QUALIFY_MAX_FILE_BYTES):
                    copied.append(str(dest))
            else:
                dest = sandbox / name
                shutil.copyfile(str(entry), str(dest), follow_symlinks=False)
                copied.append(str(dest))
        return copied

    @staticmethod
    def _write_head_sample(src_path: str, dest_path: str, n_bytes: int) -> bool:
        """Copy the first ``n_bytes`` of ``src_path`` into ``dest_path`` (a head
        sample of an oversize file — header + many rows, enough to inspect axes
        and sample IDs). Returns True on success."""
        try:
            with open(src_path, "rb") as fin, open(dest_path, "wb") as fout:
                remaining = n_bytes
                while remaining > 0:
                    chunk = fin.read(min(1024 * 1024, remaining))
                    if not chunk:
                        break
                    fout.write(chunk)
                    remaining -= len(chunk)
            return True
        except OSError as e:
            with contextlib.suppress(Exception):
                os.remove(dest_path)
            logger.warning("qualify: head-sample failed for %s: %s", src_path, e)
            return False

    @staticmethod
    def _convert_xlsx(src_path: str, sandbox_dir: str, base_stem: str, max_bytes: int) -> List[str]:
        """Convert each sheet of an .xlsx to a CSV in the sandbox (read-only
        streaming; total output capped at ``max_bytes``). Returns the written
        CSV paths; missing openpyxl or a parse error → [] (logged, never raises)."""
        try:
            import csv as _csv
            from openpyxl import load_workbook
        except Exception as e:
            logger.warning("qualify: xlsx support unavailable (openpyxl): %s", e)
            return []
        try:
            wb = load_workbook(src_path, read_only=True, data_only=True)
        except Exception as e:
            logger.warning("qualify: failed to open xlsx %s: %s", src_path, e)
            return []
        written: List[str] = []
        total = 0
        try:
            for sheet in wb.worksheets:
                safe = re.sub(r"[^A-Za-z0-9._-]", "_", sheet.title)[:40] or "sheet"
                dest = os.path.join(sandbox_dir, f"{base_stem}__{safe}.csv")
                overflow = False
                try:
                    with open(dest, "w", newline="", encoding="utf-8") as fh:
                        w = _csv.writer(fh)
                        for row in sheet.iter_rows(values_only=True):
                            cells = ["" if v is None else str(v) for v in row]
                            w.writerow(cells)
                            total += sum(len(c) for c in cells) + len(cells)
                            if total > max_bytes:
                                overflow = True
                                break
                    written.append(dest)
                except Exception as e:
                    with contextlib.suppress(Exception):
                        os.remove(dest)
                    logger.warning("qualify: xlsx sheet convert failed %s: %s", dest, e)
                if overflow:
                    logger.warning("qualify: xlsx %s truncated at size cap", src_path)
                    break
        finally:
            with contextlib.suppress(Exception):
                wb.close()
        return written

    @classmethod
    def _extract_archive(cls, src_path: str, sandbox_dir: str, base_stem: str) -> List[str]:
        """Safely extract allowlisted members of a .zip/.tar(.gz/.tgz) into the
        sandbox (datasets like GDSC/PharmacoGx ship tables inside archives).
        Members are filtered by the same allowlist (with gz/xlsx conversion);
        code/binaries and nested archives are skipped. Guards: zip-slip / path
        traversal, per-member + total decompressed-size caps, and a max file
        count. Returns the written sandbox paths; never raises."""
        import tarfile
        import tempfile
        import zipfile

        sandbox = Path(sandbox_dir)
        written: List[str] = []
        state = {"total": 0, "count": 0}
        tmp = Path(tempfile.mkdtemp(prefix="qualify-arch-"))
        try:
            if src_path.lower().endswith(".zip"):
                try:
                    with zipfile.ZipFile(src_path) as zf:
                        for info in zf.infolist():
                            if info.is_dir():
                                continue
                            if not cls._handle_archive_member(
                                info.filename, (lambda inf=info: zf.open(inf, "r")),
                                sandbox, base_stem, tmp, state, written,
                            ):
                                break
                except Exception as e:
                    # Truncated/corrupt zips raise BadZipFile, OSError, EOFError,
                    # or zlib.error mid-iteration. Honor the "never raises"
                    # contract: keep members extracted before the break.
                    logger.warning("qualify: bad zip %s: %s", src_path, e)
            else:
                try:
                    with tarfile.open(src_path, "r:*") as tf:
                        for m in tf:
                            if not m.isfile() or m.issym() or m.islnk():
                                continue
                            if not cls._handle_archive_member(
                                m.name, (lambda mm=m: tf.extractfile(mm)),
                                sandbox, base_stem, tmp, state, written,
                            ):
                                break
                except Exception as e:
                    # Truncated/corrupt tars raise TarError, OSError, EOFError,
                    # or zlib.error mid-iteration. Honor the "never raises"
                    # contract: keep members extracted before the break.
                    logger.warning("qualify: bad tar %s: %s", src_path, e)
        finally:
            shutil.rmtree(str(tmp), ignore_errors=True)
        return written

    @classmethod
    def _handle_archive_member(cls, member_name, open_fn, sandbox, base_stem, tmp, state, written) -> bool:
        """Process one archive member. Returns False to STOP iterating (caps hit),
        True to continue (member processed or skipped)."""
        if state["count"] >= QUALIFY_ARCHIVE_MAX_FILES:
            return False
        if state["total"] >= QUALIFY_ARCHIVE_MAX_TOTAL_BYTES:
            return False
        # zip-slip / traversal / absolute-path guard.
        norm = os.path.normpath(member_name or "")
        if not norm or norm.startswith((os.sep, "/")) or norm == ".." or norm.startswith(".." + os.sep) \
                or ".." in norm.replace("\\", "/").split("/"):
            return True
        base = os.path.basename(member_name)
        if not base:
            return True
        ext = os.path.splitext(base)[1].lower()
        is_member_gz = (
            ext == ".gz" and not base.lower().endswith(".tar.gz")
            and os.path.splitext(base[:-3])[1].lower() in QUALIFY_ALLOWED_EXTENSIONS
        )
        is_member_xlsx = ext == ".xlsx"
        is_member_direct = ext in QUALIFY_ALLOWED_EXTENSIONS
        if not (is_member_direct or is_member_gz or is_member_xlsx):
            return True  # code / binary / nested archive → skip
        cap = min(QUALIFY_MAX_FILE_BYTES, QUALIFY_ARCHIVE_MAX_TOTAL_BYTES - state["total"])
        raw = tmp / f"m{state['count']}_{base}"
        wrote = 0
        try:
            fh = open_fn()
            if fh is None:
                return True
            with fh, open(raw, "wb") as out:
                while True:
                    chunk = fh.read(1024 * 1024)
                    if not chunk:
                        break
                    wrote += len(chunk)
                    if wrote > cap:
                        raise ValueError("archive member exceeds size cap")
                    out.write(chunk)
        except Exception as e:
            with contextlib.suppress(Exception):
                os.remove(raw)
            logger.warning("qualify: archive member %s failed: %s", member_name, e)
            return True
        state["total"] += wrote
        state["count"] += 1
        # Flatten the member path into a sandbox-safe, collision-resistant name.
        flat = (base_stem + "__" + member_name.replace("\\", "/").replace("/", "__"))[:180]
        if is_member_direct:
            dest = sandbox / flat
            with contextlib.suppress(Exception):
                shutil.move(str(raw), str(dest))
                written.append(str(dest))
        elif is_member_gz:
            dest = sandbox / (flat[:-3] if flat.lower().endswith(".gz") else flat + ".txt")
            if cls._safe_gunzip(str(raw), str(dest), QUALIFY_MAX_FILE_BYTES):
                written.append(str(dest))
        elif is_member_xlsx:
            stem = flat[:-5] if flat.lower().endswith(".xlsx") else flat
            written.extend(cls._convert_xlsx(str(raw), str(sandbox), stem, QUALIFY_MAX_FILE_BYTES))
        return True

    @staticmethod
    def _safe_gunzip(src_path: str, dest_path: str, max_bytes: int) -> bool:
        """Decompress a single-member gzip into ``dest_path``, bounded to
        ``max_bytes`` (head sample). Uses an incremental zlib decompressor so a
        TRUNCATED gzip — e.g. an upstream ``__SAMPLE__`` head-only download — still
        yields its decoded head instead of being discarded (``gzip.read`` raises
        without returning the partial chunk; ``decompressobj`` returns output as
        it goes and simply stops at the truncation). Returns True if any head was
        written or the stream decoded cleanly; False only if nothing is readable."""
        written = 0
        try:
            decomp = zlib.decompressobj(31)  # 31 = gzip header + trailer
            with open(src_path, "rb") as fin, open(dest_path, "wb") as fout:
                while written < max_bytes:
                    comp = fin.read(1024 * 1024)
                    if not comp:
                        break
                    try:
                        out = decomp.decompress(comp, max_bytes - written)
                    except zlib.error as e:
                        logger.warning(
                            "qualify: %s gunzip stopped mid-stream (%s); kept %d-byte head",
                            src_path, e, written,
                        )
                        break
                    if out:
                        fout.write(out)
                        written += len(out)
            if written > 0:
                return True
            with contextlib.suppress(Exception):
                os.remove(dest_path)
            logger.warning("qualify: gunzip produced no output for %s", src_path)
            return False
        except Exception as e:
            if written > 0:
                return True
            with contextlib.suppress(Exception):
                os.remove(dest_path)
            logger.warning("qualify: gunzip failed for %s: %s", src_path, e)
            return False

    # ----------------------------------------------------------- claude runner

    async def _claude_health_check(self) -> bool:
        """Trivial ``claude -p`` probe. Unavailable (binary missing / non-zero /
        timeout) → False so the whole stage records ``error`` without blocking
        already-completed stages."""
        try:
            proc = await asyncio.create_subprocess_exec(
                *_claude_argv(QUALIFY_CLAUDE_MODEL, "-p", "Reply with the single word: ok"),
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                start_new_session=True,
            )
        except (FileNotFoundError, OSError):
            return False
        try:
            await asyncio.wait_for(
                proc.communicate(), timeout=QUALIFY_HEALTHCHECK_TIMEOUT_SECONDS
            )
        except (asyncio.TimeoutError, Exception):
            await self._kill_process_group(proc)
            return False
        return proc.returncode == 0

    async def _run_claude_on_dir(self, sandbox_dir: str, *, known_db: bool = False) -> Tuple[str, str, Dict]:
        """Judge the unit with the model ladder: run the cheaper PRIMARY model
        first, and escalate to the stronger model when the primary cannot give a
        usable verdict — either a technical ``error`` (unparseable/non-zero/
        timeout) or a self-declared ``escalate`` ("I can't confidently apply the
        contract"). Returns (verdict, reason, meta): a committed verdict
        (strict/loose/fail) with its decisive REASON, or ``error`` with the last
        failure cause if no model commits. ``meta`` aggregates token/cost/duration
        across every ladder attempt and names the committing model. Every
        non-commit is logged."""
        reference_gene_count = self._materialize_reference_genes(sandbox_dir)
        prompt = self.render_validation_prompt(
            sandbox_dir, reference_gene_count=reference_gene_count, known_db=known_db
        )
        name = os.path.basename(sandbox_dir.rstrip("/")) or sandbox_dir
        ladder = _qualify_model_ladder()
        agg = _empty_meta()
        last_diag = ""
        for i, model in enumerate(ladder):
            verdict, diag, meta = await self._run_one_model(sandbox_dir, prompt, model, name)
            _accumulate_meta(agg, meta)
            last_diag = diag
            if verdict in _COMMITTED_VERDICTS:
                agg["model"] = (meta or {}).get("model") or model
                return verdict, diag, agg
            is_last = i == len(ladder) - 1
            if is_last:
                agg["model"] = (meta or {}).get("model") or model
            pipeline_log(
                f"qualify: model {model!r} → {verdict} dir={name}: {diag}; "
                f"{'no more models → error' if is_last else 'escalating to next model'}",
                stage="qualify", team=self.team_name, project_id=self.project_id,
                level=logging.ERROR if is_last else logging.WARNING,
            )
        return "error", last_diag, agg

    async def _run_one_model(
        self, sandbox_dir: str, prompt: str, model: str, name: str
    ) -> Tuple[str, str, Dict]:
        """Run one model with the transient-retry loop. Returns (verdict, diag,
        meta) where verdict is a committed verdict, ``escalate``, or ``error``;
        meta is the last attempt's telemetry."""
        meta: Dict = {}
        for attempt in range(1, QUALIFY_CLAUDE_MAX_ATTEMPTS + 1):
            verdict, diag, retryable, meta = await self._invoke_claude_once(
                sandbox_dir, prompt, model
            )
            if verdict != "error":
                return verdict, diag, meta  # committed verdict or 'escalate'
            if attempt >= QUALIFY_CLAUDE_MAX_ATTEMPTS or not retryable:
                return "error", diag, meta
            pipeline_log(
                f"qualify: model {model!r} transient error (attempt {attempt}/"
                f"{QUALIFY_CLAUDE_MAX_ATTEMPTS}) dir={name}: {diag}; retrying",
                stage="qualify", team=self.team_name, project_id=self.project_id,
                level=logging.WARNING,
            )
        return "error", "", meta

    async def _invoke_claude_once(
        self, sandbox_dir: str, prompt: str, model: str = ""
    ) -> Tuple[str, str, bool, Dict]:
        """One ``claude -p --output-format json`` run with the given model.
        Returns (verdict, diagnostic, retryable, meta). verdict may be a committed
        verdict, ``escalate``, or ``error``; ``diagnostic`` carries the decisive
        ``REASON`` for a verdict or the failure cause for an error; ``meta`` holds
        the run's model / token / cost / duration telemetry (empty on spawn/parse
        failure); retryable flags transient failures (spawn / communicate /
        non-zero exit) vs deterministic ones (timeout / unparseable verdict)."""
        try:
            proc = await asyncio.create_subprocess_exec(
                *_claude_argv(model, "-p", prompt, "--output-format", "json"),
                cwd=sandbox_dir,
                stdout=asyncio.subprocess.PIPE,
                stderr=asyncio.subprocess.PIPE,
                start_new_session=True,
            )
        except (FileNotFoundError, OSError) as e:
            return "error", f"spawn failed: {type(e).__name__}: {e}", True, {}

        try:
            stdout, stderr = await asyncio.wait_for(
                proc.communicate(), timeout=QUALIFY_CLAUDE_TIMEOUT_SECONDS
            )
        except asyncio.TimeoutError:
            await self._kill_process_group(proc)
            return "error", f"timeout after {QUALIFY_CLAUDE_TIMEOUT_SECONDS:.0f}s", False, {}
        except Exception as e:
            await self._kill_process_group(proc)
            return "error", f"communicate failed: {type(e).__name__}: {e}", True, {}

        if proc.returncode != 0:
            err_tail = _to_text(stderr).strip()[-300:]
            return "error", f"non-zero exit rc={proc.returncode} stderr_tail={err_tail!r}", True, {}

        try:
            envelope = json.loads(_to_text(stdout))
        except (ValueError, TypeError):
            out_tail = " ".join(_to_text(stdout).split())[-300:]
            return "error", f"unparseable json envelope output_tail={out_tail!r}", False, {}

        meta = _extract_meta(envelope, model)
        if envelope.get("is_error") or envelope.get("subtype") != "success":
            return "error", f"claude reported error: subtype={envelope.get('subtype')!r}", True, meta

        result_text = _to_text(envelope.get("result"))
        verdict = self.parse_verdict(result_text)
        if verdict == "error":
            out_tail = " ".join(result_text.split())[-300:]
            return "error", f"unparseable verdict (rc=0) result_tail={out_tail!r}", False, meta
        return verdict, self.parse_reason(result_text), False, meta

    @staticmethod
    async def _kill_process_group(proc) -> None:
        """Kill the subprocess's whole process group (it was started with
        ``start_new_session=True``); fall back to killing the process. Always
        reap the child via ``proc.wait()`` afterwards to avoid a zombie and an
        unclosed asyncio transport (mirrors ``download/local_cdp_runtime``)."""
        killed = False
        with contextlib.suppress(Exception):
            pgid = os.getpgid(proc.pid)
            os.killpg(pgid, signal.SIGKILL)
            killed = True
        if not killed:
            with contextlib.suppress(Exception):
                proc.kill()
        with contextlib.suppress(Exception):
            await asyncio.wait_for(proc.wait(), timeout=5.0)

    # ------------------------------------------------------------ materialize

    def _materialize_unit(self, members: List[Dict], sandbox_dir: str) -> List[str]:
        """Stage a dataset unit into the sandbox, preferring each member's LOCAL
        copy (the download stage keeps local files until qualify runs)
        and downloading from remote storage only when the local file is gone.
        The hostile trust boundary (symlink/`..`/ext/size + gz/xlsx handling)
        is applied by ``_copy_into_sandbox`` on the staged dir."""
        import tempfile

        staged = tempfile.mkdtemp(prefix="qualify-stage-")
        storage = None
        got = False
        try:
            for row in members:
                fp = row.get("file_path")
                file_name = row.get("file_name") or (os.path.basename(fp) if fp else None)
                if not file_name:
                    continue
                dest = os.path.join(staged, os.path.basename(file_name))
                if fp and os.path.isfile(fp):
                    try:
                        shutil.copyfile(fp, dest)
                        got = True
                        continue
                    except OSError as e:
                        logger.warning("qualify: failed to copy local %s: %s", fp, e)
                key = row.get("oss_key") or row.get("s3_key")
                if not key:
                    continue
                if storage is None:
                    from storage.factory import get_storage
                    storage = get_storage()
                try:
                    storage.download(key, dest)
                    got = True
                except Exception as e:
                    logger.warning("qualify: failed to materialize %s: %s", key, e)
            if not got:
                return []
            return self._copy_into_sandbox(staged, sandbox_dir)
        finally:
            shutil.rmtree(staged, ignore_errors=True)

    def _cleanup_unit_local(self, members: List[Dict]) -> None:
        """Deferred disk hygiene: after a unit is qualified, delete its LOCAL
        copies that are safely in remote storage (have oss_key/s3_key). Skipped
        for the ``local`` backend (the local file is canonical) and when the
        STORAGE_DELETE_LOCAL_AFTER_UPLOAD switch is off. Local-only members
        (no remote key, e.g. embedded extractions) are kept. Best-effort."""
        if self.storage_backend not in ("oss", "s3"):
            return
        raw = os.getenv("STORAGE_DELETE_LOCAL_AFTER_UPLOAD", "true").strip().lower()
        if raw in ("false", "0", "no", "off", ""):
            return
        for row in members:
            if not (row.get("oss_key") or row.get("s3_key")):
                continue
            fp = row.get("file_path")
            if fp and os.path.isfile(fp):
                try:
                    os.remove(fp)
                except OSError as e:
                    logger.warning("qualify: could not remove local copy %s: %s", fp, e)

    def _sweep_local_after_qualify(self) -> None:
        """Fallback cleanup when qualify could not run (health-check fail): the
        download stage deferred deletion to qualify, so free the deferred local
        copies (they are already in remote storage) to avoid a disk leak."""
        if self.storage_backend not in ("oss", "s3"):
            return
        try:
            rows = self._get_dao().get_dataset_artifacts_for_qualify(self.project_id)
        except Exception:
            return
        self._cleanup_unit_local(rows)

    # -------------------------------------------------------------------- run

    async def run(self) -> Optional[Dict[str, int]]:
        import tempfile

        update_stage("Stage 4 - Qualifying Datasets")
        push_message("Qualify stage started", team_name=self.team_name)

        if not await self._claude_health_check():
            msg = (
                "Claude Code (`claude`) is unavailable — qualify stage cannot run. "
                "Install Claude Code and ensure non-interactive auth is configured "
                "(claude -p) on this host. Already-completed stages are unaffected."
            )
            pipeline_log(
                msg,
                stage="qualify",
                team=self.team_name,
                project_id=self.project_id,
                level=logging.ERROR,
            )
            push_message(msg, team_name=self.team_name, message_type="error")
            # Download deferred local deletion to us; free the deferred copies
            # (already in remote storage) so a disabled/broken qualify doesn't
            # leak disk.
            self._sweep_local_after_qualify()
            # None is the sentinel for "stage could not run" (health-check
            # failure) vs a dict of per-dataset counts when it did run. The
            # caller fails the stage only on None, never on per-dataset errors.
            return None

        dao = self._get_dao()
        rows = dao.get_dataset_artifacts_for_qualify(self.project_id)
        groups = self.group_dataset_units(rows)
        if not groups:
            push_message("No datasets to qualify", team_name=self.team_name)
            update_progress(1.0)
            return {"strict": 0, "loose": 0, "fail": 0, "error": 0}

        counts = {"strict": 0, "loose": 0, "fail": 0, "error": 0}
        total = len(groups)
        for idx, (unit_dir, members) in enumerate(groups.items()):
            member_ids = [r["id"] for r in members if r.get("id") is not None]

            # Idempotency: a unit already holding a settled verdict (deterministic
            # per the model ladder) is not re-run — skip the claude call so a
            # re-invocation costs nothing. Only pending/processing/error reprocess.
            settled = {r.get("qualification_status") for r in members}
            if len(settled) == 1 and next(iter(settled)) in ("strict", "loose", "fail"):
                verdict = next(iter(settled))
                counts[verdict] = counts.get(verdict, 0) + 1
                push_message(
                    f"Skipped {os.path.basename(unit_dir) or unit_dir}: already {verdict}",
                    team_name=self.team_name,
                )
                update_progress((idx + 1) / total)
                continue

            dao.set_qualification_status(self.project_id, member_ids, "processing")

            sandbox = tempfile.mkdtemp(prefix="qualify-sandbox-")
            try:
                copied = self._materialize_unit(members, sandbox)
                if not copied:
                    verdict, reason, meta = "error", "no files materialized from dataset unit", {}
                else:
                    known_db = os.path.basename(unit_dir.rstrip("/")).startswith("_knowndb_")
                    verdict, reason, meta = await self._run_claude_on_dir(sandbox, known_db=known_db)
            except Exception as e:
                # Per-unit isolation: one bad dataset (corrupt archive, OOM,
                # claude crash) must NOT abort qualify for the remaining units.
                logger.exception("qualify: unit %s failed", unit_dir)
                verdict, reason, meta = (
                    "error", f"qualify unit crashed: {type(e).__name__}: {e}"[:300], {}
                )
            finally:
                shutil.rmtree(sandbox, ignore_errors=True)

            dao.set_qualification_status(
                self.project_id, member_ids, verdict, reason=reason
            )
            # Cost/token/duration telemetry → stage_costs (per-paper aggregation).
            dao.record_stage_cost(
                self.project_id, os.path.basename(unit_dir.rstrip("/")), "qualify", meta
            )
            # Deferred disk hygiene: drop this unit's local copies now that it's
            # validated (only those already safe in remote storage).
            self._cleanup_unit_local(members)
            counts[verdict] = counts.get(verdict, 0) + 1
            push_message(
                f"Qualified {os.path.basename(unit_dir) or unit_dir}: {verdict}",
                team_name=self.team_name,
            )
            update_progress((idx + 1) / total)

        pipeline_log(
            f"qualify complete: {counts}",
            stage="qualify",
            team=self.team_name,
            project_id=self.project_id,
        )
        return counts

    async def cleanup(self):
        self._dao = None


def _to_text(data) -> str:
    if isinstance(data, bytes):
        return data.decode("utf-8", errors="replace")
    return data or ""
